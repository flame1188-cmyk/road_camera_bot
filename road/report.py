"""
Модуль генерации отчётов (Excel) по результатам оценки участка дороги.
"""

from __future__ import annotations

import io
import json
import logging
from datetime import datetime
from typing import Any

logger = logging.getLogger(__name__)


def generate_excel_report(
    lat: float, lon: float, address: str,
    vlm_result: dict[str, Any], osm_data: dict[str, Any] | None = None,
    gibdd_nearby: list[dict] | None = None,
    panorama_images: list[dict] | None = None,
    narodnaya_map_bytes: bytes | None = None,
) -> bytes:
    """Генерирует Excel-файл с результатами оценки участка дороги."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        logger.error(f"openpyxl не установлен: {e}")
        return b""

    if not vlm_result:
        logger.warning("Excel: vlm_result пустой, генерация с минимальными данными")
        vlm_result = {"expediency": {"efficiency_score": 0}, "infrastructure": {},
                       "road_objects": {}, "technical_feasibility": {}, "visual_notes": ""}

    wb = Workbook()
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # --- Лист 1: Сводка ---
    ws1 = wb.active
    ws1.title = "Сводная оценка"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "ЭКСПЕРТНАЯ ОЦЕНКА УЧАСТКА ДОРОГИ"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center")

    row = 3
    for label, value in [("Дата оценки", datetime.now().strftime("%d.%m.%Y %H:%M")), ("Широта", str(lat)), ("Долгота", str(lon)), ("Адрес", address)]:
        ws1[f"A{row}"] = label
        ws1[f"A{row}"].font = header_font
        ws1[f"B{row}"] = value
        row += 1

    row += 1
    ws1[f"A{row}"] = "ОЦЕНКА ЦЕЛЕСООБРАЗНОСТИ"
    ws1[f"A{row}"].font = header_font
    row += 1
    exp = vlm_result.get("expediency", {})
    score = exp.get("efficiency_score", 0)
    ws1[f"A{row}"] = "Оценка эффективности"
    ws1[f"B{row}"] = f"{score}/10"
    ws1[f"B{row}"].fill = green_fill if score >= 7 else (yellow_fill if score >= 4 else red_fill)
    row += 1

    row += 1
    ws1[f"A{row}"] = "ИНФРАСТРУКТУРА"
    ws1[f"A{row}"].font = header_font
    row += 1
    infra = vlm_result.get("infrastructure", {})
    for label, value in [("Опоры освещения", infra.get("lighting_poles", "не видно")), ("Количество опор", str(infra.get("pole_count", 0))),
                         ("Провода", infra.get("wires_visible", "не видно")), ("Тип дороги", infra.get("road_type", "?")),
                         ("Полосы", str(infra.get("lane_count", "?"))), ("Разделительная полоса", infra.get("median", "нет"))]:
        ws1[f"A{row}"] = f"  {label}"
        ws1[f"B{row}"] = value
        row += 1

    row += 1
    ws1[f"A{row}"] = "ДОРОЖНЫЕ ОБЪЕКТЫ"
    ws1[f"A{row}"].font = header_font
    row += 1
    objects = vlm_result.get("road_objects", {})
    ws1[f"A{row}"] = "  Знаки"
    ws1[f"B{row}"] = ", ".join(str(s) for s in objects.get("signs", [])) or "не обнаружены"
    row += 1
    cw = objects.get("crosswalk", {})
    ws1[f"A{row}"] = "  Пешеходный переход"
    ws1[f"B{row}"] = f"{'есть' if cw.get('present') else 'нет'} ({cw.get('type', '-')})"
    row += 1
    ws1[f"A{row}"] = "  Светофор"
    ws1[f"B{row}"] = "есть" if objects.get("traffic_light") else "нет"
    row += 1

    row += 1
    ws1[f"A{row}"] = "НАРУШЕНИЯ"
    ws1[f"A{row}"].font = header_font
    row += 1
    for v in exp.get("possible_violations", []):
        ws1[f"A{row}"] = f"  • {v}"
        row += 1
    ws1[f"A{row}"] = "  Рекомендуемый тип"
    ws1[f"B{row}"] = exp.get("recommended_type", "?")
    row += 1

    row += 1
    ws1[f"A{row}"] = "ТЕХНИЧЕСКАЯ ВОЗМОЖНОСТЬ"
    ws1[f"A{row}"].font = header_font
    row += 1
    tech = vlm_result.get("technical_feasibility", {})
    for label, value in [("Питание", tech.get("power_supply", "?")), ("Установка на опору", "да" if tech.get("install_on_existing_pole") else "нет"),
                         ("Фундамент", "необходим" if tech.get("foundation_needed") else "не требуется"), ("Обзорность", tech.get("visibility_assessment", "?"))]:
        ws1[f"A{row}"] = f"  {label}"
        ws1[f"B{row}"] = value
        row += 1

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 60

    # --- Лист 2: OSM ---
    if osm_data:
        ws2 = wb.create_sheet("Данные OSM")
        road = osm_data.get("road_info", {})
        row2 = 1
        ws2[f"A{row2}"] = "ДАННЫЕ OPENSTREETMAP"
        ws2[f"A{row2}"].font = title_font
        row2 += 2
        for label, value in [("Тип дороги", road.get("road_type", "")), ("Название", road.get("road_name", "")),
                             ("Полосы", road.get("lanes", "")), ("Покрытие", road.get("surface", "")),
                             ("Опоры (200м)", osm_data.get("streetlamp_count", 0)),
                             ("Переходы (100м)", osm_data.get("crossing_count", 0)),
                             ("Светофоры (100м)", osm_data.get("traffic_signal_count", 0)),
                             ("Остановки (300м)", osm_data.get("bus_stop_count", 0)),
                             ("Школы (300м)", osm_data.get("school_count", 0)),
                             ("Детсады (300м)", osm_data.get("kindergarten_count", 0))]:
            ws2[f"A{row2}"] = label
            ws2[f"B{row2}"] = str(value)
            row2 += 1
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 50

    # --- Лист 3: ДТП ---
    if gibdd_nearby:
        ws3 = wb.create_sheet("ДТП поблизости")
        ws3[f"A1"] = "ДТП В РАДИУСЕ 500 М"
        ws3[f"A1"].font = title_font
        headers = ["Дата", "Время", "Вид ДТП", "Погибло", "Ранено", "Улица", "Широта", "Долгота"]
        for ci, h in enumerate(headers, 1):
            cell = ws3.cell(row=2, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for ri, acc in enumerate(gibdd_nearby, 3):
            for ci, key in enumerate(["date_dtp", "time", "dtpv", "pog", "ran", "street", "coord_w", "coord_l"], 1):
                cell = ws3.cell(row=ri, column=ci, value=str(acc.get(key, "")))
                cell.font = normal_font
                cell.border = thin_border

    # --- Лист 4: Панорамы ---
    if panorama_images or narodnaya_map_bytes:
        try:
            from openpyxl.drawing.image import Image as XlImage
            ws_pan = wb.create_sheet("Панорамы")
            ws_pan[f"A1"] = "ФОТОГРАФИИ УЧАСТКА"
            ws_pan[f"A1"].font = title_font
            ws_pan[f"A2"] = "Использованы для VLM-анализа"
            row_pan = 4

            # Народная карта (первой — она содержит скоростные ограничения)
            if narodnaya_map_bytes:
                ws_pan.cell(row=row_pan, column=1, value="Народная карта (скоростные режимы)").font = header_font
                row_pan += 1
                img_stream = io.BytesIO(narodnaya_map_bytes)
                img = XlImage(img_stream)
                img.width = 640
                img.height = 360
                ws_pan.add_image(img, f"A{row_pan}")
                row_pan += 20

            # Панорамы
            for pano in panorama_images:
                heading = pano.get("heading", 0)
                pano_bytes = pano.get("bytes")
                if not pano_bytes:
                    continue
                ws_pan.cell(row=row_pan, column=1, value=f"Панорама {int(heading)}°").font = header_font
                row_pan += 1
                img_stream = io.BytesIO(pano_bytes)
                img = XlImage(img_stream)
                img.width = 640
                img.height = 360
                ws_pan.add_image(img, f"A{row_pan}")
                row_pan += 20  # отступ для следующего фото
            ws_pan.column_dimensions["A"].width = 15
        except Exception as e:
            logger.warning(f"Панорамы в Excel: {e}")

    # --- Лист 5: JSON ---
    ws4 = wb.create_sheet("Полные данные")
    ws4["A1"] = "Полный JSON VLM"
    ws4["A1"].font = title_font
    ws4["A2"] = json.dumps(vlm_result, ensure_ascii=False, indent=2)
    ws4.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    try:
        wb.save(buf)
        buf.seek(0)
        result = buf.getvalue()
        logger.info(f"Excel: сгенерирован {len(result)} байт, {len(wb.sheetnames)} листов")
        return result
    except Exception as e:
        logger.error(f"Excel: ошибка сохранения wb — {e}")
        return b""


def get_report_filename(lat: float, lon: float) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"road_assessment_{lat}_{lon}_{ts}.xlsx"
