"""
Генерация Excel-файлов на основе данных ДТП:

  1. dtp_cards.xlsx     — одна строка = одно ДТП (все поля карточки)
  2. dtp_uch.xlsx       — одна строка = один участник ДТП
  3. dtp_analytics.xlsx — аналитика: сравнение периодов (цветовое кодирование)
  4. dtp_conc.xlsx      — очаги ДТП: сводка + детализация (2 листа)
"""

from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


# ========================
# Стили
# ========================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CELL_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ========================
# Вспомогательные функции
# ========================

def _apply_header_style(ws, col_count: int) -> None:
    """Применяет стили к строке заголовков."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_data_styles(ws, row_count: int, col_count: int) -> None:
    """Применяет стили к ячейкам с данными."""
    for row_idx in range(2, row_count + 2):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, col_count: int, max_width: int = 40) -> None:
    """Автоподбор ширины колонок с ограничением."""
    for col_idx in range(1, col_count + 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))

        check_rows = min(ws.max_row, 51)
        for row_idx in range(2, check_rows):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                cell_len = len(str(cell_val))
                if cell_len > max_len:
                    max_len = cell_len

        adjusted_width = min(max_len + 3, max_width)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 8)


def _create_workbook(
    column_names: list[str],
    data_rows: list[dict[str, str]],
) -> Workbook:
    """
    Создаёт объект Workbook с заголовками и данными.

    Args:
        column_names: Список названий колонок (порядок важен)
        data_rows: Список словарей {название_колонки: значение}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    # Заголовки
    for col_idx, col_name in enumerate(column_names, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Данные
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    col_count = len(column_names)
    row_count = len(data_rows)

    _apply_header_style(ws, col_count)
    _apply_data_styles(ws, row_count, col_count)
    _auto_width(ws, col_count)

    ws.freeze_panes = "A2"

    if row_count > 0:
        ws.auto_filter.ref = ws.dimensions

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Сериализует Workbook в байты xlsx-файла."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ========================
# Публичные функции
# ========================

def generate_file1(file1_data: list[dict[str, str]]) -> bytes:
    """
    Генерирует Файл 1: одна строка = одно ДТП.

    Args:
        file1_data: Данные от gibdd.parser.build_file1_data()

    Returns:
        Байты xlsx-файла
    """
    from gibdd.parser import get_file1_column_names

    column_names = get_file1_column_names()
    wb = _create_workbook(column_names, file1_data)
    return workbook_to_bytes(wb)


def generate_file2(file2_data: list[dict[str, str]]) -> bytes:
    """
    Генерирует Файл 2: одна строка = один участник.

    Args:
        file2_data: Данные от gibdd.parser.build_file2_data()

    Returns:
        Байты xlsx-файла
    """
    from gibdd.parser import get_file2_column_names

    column_names = get_file2_column_names()
    wb = _create_workbook(column_names, file2_data)
    return workbook_to_bytes(wb)


def generate_both_files(
    file1_data: list[dict[str, str]],
    file2_data: list[dict[str, str]],
) -> tuple[bytes, bytes]:
    """
    Генерирует оба Excel-файла.

    Returns:
        (file1_bytes, file2_bytes)
    """
    logger.info(f"Генерация Excel: Файл 1 — {len(file1_data)} ДТП, Файл 2 — {len(file2_data)} участников")
    file1_bytes = generate_file1(file1_data)
    file2_bytes = generate_file2(file2_data)
    logger.info(f"Файл 1: {len(file1_bytes)} байт, Файл 2: {len(file2_bytes)} байт")
    return file1_bytes, file2_bytes


# ========================
# Стили для аналитики
# ========================

SECTION_FONT = Font(bold=True, size=11)
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def generate_analytics_file(
    analytics_data: list[dict[str, str]],
    column_names: list[str],
) -> bytes:
    """
    Генерирует Excel-файл с аналитикой (сравнение периодов).
    Цветовое кодирование изменений: зелёный = снижение, красный = рост.

    Args:
        analytics_data: Данные от gibdd.analytics.build_analytics_excel_data()
        column_names: Названия колонок от gibdd.analytics.get_analytics_column_names()

    Returns:
        Байты xlsx-файла
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитика"

    col_count = len(column_names)

    # Заголовки
    for col_idx, col_name in enumerate(column_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Данные
    for row_idx, row_data in enumerate(analytics_data, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # Выделяем разделительные строки (заголовки секций)
        indicator = row_data.get("Показатель", "")
        if indicator and indicator == indicator.upper() and indicator.strip():
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = SECTION_FONT

        # Цветовое кодирование колонки "Изменение, %"
        change_cell = ws.cell(row=row_idx, column=4)
        change_val = row_data.get("Изменение, %", "")
        if isinstance(change_val, (int, float)) and change_val != 0:
            if change_val > 0:
                change_cell.fill = NEGATIVE_FILL  # Рост показателя = красный
            elif change_val < 0:
                change_cell.fill = POSITIVE_FILL  # Снижение показателя = зелёный

    # Ширина колонок
    ws.column_dimensions["A"].width = 35
    for col_idx in range(2, col_count + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = "A2"

    return workbook_to_bytes(wb)


# ========================
# Очаги концентрации ДТП
# ========================

ZONE_NP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ZONE_NONP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def generate_concentration_file(
    concentration_data: list[dict[str, str]],
    column_names: list[str],
    detail_data: list[dict[str, str]] | None = None,
    detail_columns: list[str] | None = None,
) -> bytes:
    """
    Генерирует Excel-файл с очагами концентрации ДТП.
    Лист 1 — сводка очагов, Лист 2 — детализация ДТП.

    Цветовое кодирование по типу зоны:
      - НП — жёлтый
      - Вне НП — голубой

    Args:
        concentration_data: Данные от gibdd.concentration.build_concentration_excel_data()
        column_names: Названия колонок от gibdd.concentration.get_concentration_column_names()
        detail_data: Данные от gibdd.concentration.build_concentration_detail_data() (опционально)
        detail_columns: Названия колонок от gibdd.concentration.get_detail_column_names() (опционально)

    Returns:
        Байты xlsx-файла
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Очаги ДТП"

    col_count = len(column_names)

    # Заголовки
    for col_idx, col_name in enumerate(column_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Данные
    for row_idx, row_data in enumerate(concentration_data, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # Цветовое кодирование по типу зоны
        zone = row_data.get("Тип зоны", "")
        fill = None
        if zone.startswith("НП"):
            fill = ZONE_NP_FILL
        elif zone.startswith("Вне"):
            fill = ZONE_NONP_FILL

        if fill:
            for col_idx in range(1, col_count + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Ширина колонок
    col_widths = {
        "№": 6,
        "Тип зоны": 28,
        "Дорога/Улица": 35,
        "Широта": 16,
        "Долгота": 16,
        "ДТП в очаге": 10,
        "Погибло": 8,
        "Ранено": 8,
        "Доминирующий вид ДТП": 25,
        "Виды ДТП": 45,
        "Даты": 30,
    }
    for col_idx, col_name in enumerate(column_names, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    ws.freeze_panes = "A2"

    if concentration_data:
        ws.auto_filter.ref = ws.dimensions

    # --- Лист 2: Детализация ДТП в очагах ---
    if detail_data and detail_columns:
        ws2 = wb.create_sheet("Детализация ДТП")

        det_col_count = len(detail_columns)

        # Заголовки
        for col_idx, col_name in enumerate(detail_columns, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Данные
        for row_idx, row_data in enumerate(detail_data, start=2):
            for col_idx, col_name in enumerate(detail_columns, start=1):
                value = row_data.get(col_name, "")
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

        # Ширина колонок
        det_widths = {
            "№": 6,
            "Дорога": 30,
            "Дата": 14,
            "Вид ДТП": 25,
            "Погибло": 8,
            "Ранено": 8,
            "Широта": 16,
            "Долгота": 16,
            "Улица": 25,
            "НП": 20,
            "Км": 10,
        }
        for col_idx, col_name in enumerate(detail_columns, start=1):
            col_letter = ws2.cell(row=1, column=col_idx).column_letter
            ws2.column_dimensions[col_letter].width = det_widths.get(col_name, 20)

        ws2.freeze_panes = "A2"

        if detail_data:
            ws2.auto_filter.ref = ws2.dimensions

    return workbook_to_bytes(wb)
