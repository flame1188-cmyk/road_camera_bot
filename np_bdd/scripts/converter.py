"""
converter.py — конвертация Excel-таблиц пользователя в JSON-кэш НП БДД.

ОЖИДАЕМЫЙ ВХОД
==============

Пользователь кладёт в np_bdd/data/raw/ три типа файлов:

1. vehicles_<...>.xlsx  — Ктс по регионам и годам
2. plans_<...>.xlsx     — Плановые Тр из паспорта НП БДД (2023-2030)
3. history_<...>.xlsx   — (опционально) исторические погибшие 2023-2025

Имена файлов могут быть произвольными, тип определяется автоматически по
содержимому (заголовкам столбцов). См. функции detect_kind() ниже.

ВЫХОД
=====

Для каждого встреченного региона создаются (или обновляются) три файла:

- np_bdd/data/vehicles/{region_code}.json
- np_bdd/data/plans/{region_code}.json
- np_bdd/data/history/{region_code}.json

Все файлы валидируются по схемам из np_bdd/schemas/.

ИСПОЛЬЗОВАНИЕ
=============

    python np_bdd/scripts/converter.py                 # обработать все файлы в data/raw/
    python np_bdd/scripts/converter.py <file.xlsx>     # обработать один файл
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from typing import Any, Literal

import openpyxl
from jsonschema import validate

# --- Константы -------------------------------------------------------------

# --- Относительные пути ---------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent  # np_bdd/
DATA_RAW_DIR = PROJECT_ROOT / "datasets" / "raw"
DATA_HIST_DIR = PROJECT_ROOT / "datasets" / "history"
DATA_PLANS_DIR = PROJECT_ROOT / "datasets" / "plans"
DATA_VEHI_DIR = PROJECT_ROOT / "datasets" / "vehicles"
SCHEMAS_DIR = PROJECT_ROOT / "schemas"

TODAY = date.today().isoformat()

# --- Схемы -----------------------------------------------------------------


def load_schema(name: str) -> dict[str, Any]:
    path = SCHEMAS_DIR / name
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


SCHEMA_HISTORY = load_schema("history.schema.json")
SCHEMA_PLANS = load_schema("plans.schema.json")
SCHEMA_VEHICLES = load_schema("vehicles.schema.json")

# --- Определение типа файла ------------------------------------------------


def detect_kind(file_path: Path) -> Literal["vehicles", "plans", "history", "unknown"]:
    """
    Автоопределение типа файла.

    Шаг 1: по подстроке в имени файла (включая русские ключевые слова).
    Шаг 2: если шаг 1 не дал результата — по сигнатурным словам в заголовке
           первой строки листа.

    Поддерживаемые имена файлов:
      - КТС.xlsx, vehicles_kts.xlsx, kts.xlsx, _ктс_
      - Показатели ТР.xlsx, plans_tr.xlsx, passport.xlsx, _паспорт_, _план_
      - history_deaths.xlsx, _погибш_
    """
    name = file_path.name.lower()
    # 1) По имени файла.
    if "vehicles" in name or "kts" in name or "ктс" in name:
        return "vehicles"
    if "plan" in name or "passport" in name or "паспорт" in name or "план" in name or "показател" in name:
        # "план" может быть и в "history_deaths.xlsx"? Нет. Ок.
        return "plans"
    if "history" in name or "deaths" in name or "погибш" in name:
        return "history"

    # 2) По заголовкам листа.
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(values_only=True), None)
        wb.close()
        if first_row:
            headers_joined = " ".join(str(c).lower() for c in first_row if c)
            # Если в заголовках встречаются годы 2023..2030 и есть слово "регион"
            # — это либо vehicles, либо plans, либо history. Различаем по числу
            # годовых колонок: plans имеет 8 (2023..2030), vehicles/history — 4 (2023..2026).
            if "регион" in headers_joined:
                year_cols = [c for c in first_row if c is not None and str(c).isdigit() and len(str(c)) == 4]
                if len(year_cols) >= 6:
                    return "plans"
                # 4 годовые колонки — неоднозначно. По умолчанию vehicles.
                return "vehicles"
    except Exception:  # noqa: BLE001
        pass

    return "unknown"


# --- Парсеры (заглушки под реальную структуру) -----------------------------


def parse_vehicles(file_path: Path) -> list[dict[str, Any]]:
    """
    Парсит Excel с Ктс.

    Реальная структура (подтверждена файлом КТС.xlsx):

        | Регион                | Код Региона | 2023    | 2024    | 2025    | 2026    |
        |-----------------------|-------------|---------|---------|---------|---------|
        | Краснодарский край    | 1103        | 2240309 | 2289270 | 2834711 | 2834711 |
        | г. Севастополь        | 1106        | 175542  | 180241  | 147239  | 147239  |

    Особенности:
    - Колонка 0 = имя региона, колонка 1 = код (4-значный строкой).
    - Значения Ктс могут прийти строкой ('2240309') или числом — конвертируем в int.
    - Код региона — строка, сохраняем как есть.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        region_name = str(row[0]).strip() if row[0] is not None else ""
        region_code = str(row[1]).strip() if row[1] is not None else ""
        vehicles_by_year: dict[str, int] = {}
        for i, col in enumerate(header[2:], start=2):
            if col.isdigit() and len(col) == 4:
                val = row[i] if i < len(row) else None
                if val is None:
                    continue
                try:
                    # Может прийти int, float или строкой
                    vehicles_by_year[col] = int(float(str(val).replace(" ", "").replace("\xa0", "")))
                except (ValueError, TypeError):
                    continue
        if not region_code or not region_name or not vehicles_by_year:
            continue
        result.append({
            "region_code": region_code,
            "region_name": region_name,
            "vehicles_by_year": vehicles_by_year,
            "source": f"Excel: {file_path.name}",
            "loaded_at": TODAY,
        })
    return result


def parse_plans(file_path: Path) -> list[dict[str, Any]]:
    """
    Парсит Excel с плановыми Тр из паспорта НП БДД.

    Реальная структура (подтверждена файлом Показатели ТР.xlsx):

        | Регион                | Код Региона | 2023 | 2024 | ... | 2030 |
        |-----------------------|-------------|------|------|-----|------|
        | Краснодарский край    | 1103        | 3.37 | 3.17 | ... | 2.36 |

    Особенности:
    - Колонка 0 = имя региона, колонка 1 = код (4-значный).
    - Значения Тр — float (могут прийти строкой '3.37' или числом).
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        region_name = str(row[0]).strip() if row[0] is not None else ""
        region_code = str(row[1]).strip() if row[1] is not None else ""
        plan_tr: dict[str, float] = {}
        for i, col in enumerate(header[2:], start=2):
            if col.isdigit() and len(col) == 4:
                val = row[i] if i < len(row) else None
                if val is None:
                    continue
                try:
                    plan_tr[col] = float(str(val).replace(",", ".").strip())
                except (ValueError, TypeError):
                    continue
        if not region_code or not region_name or not plan_tr:
            continue
        result.append({
            "region_code": region_code,
            "region_name": region_name,
            "plan_tr": plan_tr,
            "source": f"Паспорт НП БДД, Excel: {file_path.name}",
            "loaded_at": TODAY,
        })
    return result


def parse_history(file_path: Path) -> list[dict[str, Any]]:
    """
    Парсит Excel с историческими погибшими (если пользователь не хочет
    тянуть данные с сайта ГИБДД).

    Ожидаемая структура (по аналогии с КТС.xlsx):

        | Регион                | Код Региона | 2023 | 2024 | 2025 |
        |-----------------------|-------------|------|------|------|
        | г. Севастополь        | 1106        | 18   | 16   | 14   |

    Ктс подтягивается из уже сохранённого vehicles/{region_code}.json
    и Тр считается здесь же.

    Возвращает список словарей, готовых под history.schema.json.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        region_name = str(row[0]).strip() if row[0] is not None else ""
        region_code = str(row[1]).strip() if row[1] is not None else ""

        # Загружаем Ктс из ранее сохранённого файла, если есть.
        veh_file = DATA_VEHI_DIR / f"{region_code}.json"
        vehicles_by_year: dict[str, int] = {}
        if veh_file.exists():
            with veh_file.open("r", encoding="utf-8") as fh:
                vehicles_by_year = json.load(fh).get("vehicles_by_year", {})

        years: dict[str, dict[str, Any]] = {}
        for i, col in enumerate(header[2:], start=2):
            if not (col.isdigit() and len(col) == 4):
                continue
            deaths_val = row[i] if i < len(row) else None
            if deaths_val is None:
                continue
            try:
                deaths = int(float(str(deaths_val).replace(" ", "").replace("\xa0", "")))
            except (ValueError, TypeError):
                continue
            veh = vehicles_by_year.get(col)
            if not veh:
                # Невозможно посчитать Тр без Ктс — пропускаем год.
                continue
            tr = round((deaths * 10000) / veh, 3)
            years[col] = {
                "deaths": deaths,
                "vehicles": veh,
                "tr": tr,
                "calculated_at": TODAY,
            }
        if not region_code or not region_name or not years:
            continue
        result.append({
            "region_code": region_code,
            "region_name": region_name,
            "years": years,
        })
    return result


# --- Сохранение ------------------------------------------------------------


def save_json(file_path: Path, payload: dict[str, Any], schema: dict[str, Any]) -> None:
    validate(instance=payload, schema=schema)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with file_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def process_file(file_path: Path) -> dict[str, int]:
    """Обрабатывает один Excel-файл. Возвращает счётчики."""
    kind = detect_kind(file_path)
    print(f"[converter] {file_path.name}: тип = {kind}")
    if kind == "unknown":
        print(f"[converter]   пропущен (не удалось определить тип)")
        return {"vehicles": 0, "plans": 0, "history": 0}

    if kind == "vehicles":
        items = parse_vehicles(file_path)
        for it in items:
            out = DATA_VEHI_DIR / f"{it['region_code']}.json"
            save_json(out, it, SCHEMA_VEHICLES)
        print(f"[converter]   сохранено {len(items)} regions → {DATA_VEHI_DIR}")
        return {"vehicles": len(items), "plans": 0, "history": 0}

    if kind == "plans":
        items = parse_plans(file_path)
        for it in items:
            out = DATA_PLANS_DIR / f"{it['region_code']}.json"
            save_json(out, it, SCHEMA_PLANS)
        print(f"[converter]   сохранено {len(items)} regions → {DATA_PLANS_DIR}")
        return {"vehicles": 0, "plans": len(items), "history": 0}

    if kind == "history":
        items = parse_history(file_path)
        for it in items:
            out = DATA_HIST_DIR / f"{it['region_code']}.json"
            save_json(out, it, SCHEMA_HISTORY)
        print(f"[converter]   сохранено {len(items)} regions → {DATA_HIST_DIR}")
        return {"vehicles": 0, "plans": 0, "history": len(items)}

    return {"vehicles": 0, "plans": 0, "history": 0}


def main(argv: list[str]) -> int:
    if len(argv) > 1:
        files = [Path(argv[1])]
    else:
        files = sorted(DATA_RAW_DIR.glob("*.xlsx"))

    if not files:
        print(f"[converter] нет .xlsx в {DATA_RAW_DIR}")
        print("[converter] положите туда Excel-файлы и запустите снова.")
        return 1

    total = {"vehicles": 0, "plans": 0, "history": 0}
    for f in files:
        if not f.exists():
            print(f"[converter] файл не найден: {f}")
            continue
        try:
            counts = process_file(f)
            for k in total:
                total[k] += counts[k]
        except Exception as exc:  # noqa: BLE001
            print(f"[converter] ОШИБКА при обработке {f.name}: {exc}")
            # продолжаем со следующего файла

    print(f"[converter] ИТОГО: {total}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
